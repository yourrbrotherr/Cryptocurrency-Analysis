import requests
import pandas as pd
import time
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from fpdf import FPDF

# API URL for fetching live cryptocurrency data
API_URL = "https://api.coingecko.com/api/v3/coins/markets"
PARAMS = {
    "vs_currency": "usd",
    "order": "market_cap_desc",
    "per_page": 50,
    "page": 1,
    "sparkline": False
}

def fetch_data():
    #Fetch live data for top 50 cryptocurrencies
    response = requests.get(API_URL, params=PARAMS)
    if response.status_code == 200:
        data = response.json()
        df = pd.DataFrame(data)[["name", "symbol", "current_price", "market_cap", "total_volume", "price_change_percentage_24h"]]
        df.columns = ["Name", "Symbol", "Price (USD)", "Market Cap", "24H Volume", "% Change (24H)"]
        return df
    else:
        print("Failed to fetch data.")
        return None

def perform_analysis(df):
    #Analyze the data: Top 5 by Market Cap, Average Price, and Price Change
    analysis = {}
    analysis['Top 5 by Market Cap'] = df.nlargest(5, 'Market Cap')[['Name', 'Market Cap']]
    analysis['Average Price'] = df['Price (USD)'].mean()
    analysis['Highest 24H Change'] = df.loc[df['% Change (24H)'].idxmax()][['Name', '% Change (24H)']]
    analysis['Lowest 24H Change'] = df.loc[df['% Change (24H)'].idxmin()][['Name', '% Change (24H)']]
    return analysis

def export_to_excel(df):
    #Export data to a live-updating Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "Crypto Data"

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    #Save initial Excel file
    file_name = "Live_Crypto_Data.xlsx"
    wb.save(file_name)
    print(f"Data exported to {file_name}")
    return file_name

def generate_report(analysis):
    #Generate a PDF report summarizing the analysis
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    pdf.cell(200, 10, txt="Cryptocurrency Data Analysis Report", ln=True, align="C")
    pdf.ln(10)

    #Top 5 by Market Cap
    pdf.cell(200, 10, txt="Top 5 Cryptocurrencies by Market Cap:", ln=True)
    for idx, row in analysis['Top 5 by Market Cap'].iterrows():
        pdf.cell(200, 10, txt=f"{row['Name']}: {row['Market Cap']:,}", ln=True)
    pdf.ln(10)

    #Average Price
    pdf.cell(200, 10, txt=f"Average Price of Top 50 Cryptocurrencies: ${analysis['Average Price']:.2f}", ln=True)
    pdf.ln(10)

    #Highest 24H Change
    pdf.cell(200, 10, txt=f"Highest 24H Change: {analysis['Highest 24H Change']['Name']} ({analysis['Highest 24H Change']['% Change (24H)']:.2f}%)", ln=True)
    pdf.ln(10)

    #Lowest 24H Change
    pdf.cell(200, 10, txt=f"Lowest 24H Change: {analysis['Lowest 24H Change']['Name']} ({analysis['Lowest 24H Change']['% Change (24H)']:.2f}%)", ln=True)
    pdf.ln(10)

    #Save the PDF
    file_name = "Crypto_Analysis_Report.pdf"
    pdf.output(file_name)
    print(f"Report generated: {file_name}")
    return file_name

def main():
    print("Fetching data...")
    df = fetch_data()
    if df is not None:
        print("Data fetched successfully.")

        print("Performing analysis...")
        analysis = perform_analysis(df)
        print("Analysis completed:")
        print(analysis)

        print("Exporting data to Excel...")
        export_to_excel(df)

        print("Generating analysis report...")
        generate_report(analysis)

        #Updates every 5 minutes
        print("Setting up live updates. Press Ctrl+C to stop.")
        while True:
            time.sleep(300)  # 5 minutes
            df = fetch_data()
            if df is not None:
                wb = Workbook()
                ws = wb.active
                ws.title = "Crypto Data"
                for row in dataframe_to_rows(df, index=False, header=True):
                    ws.append(row)
                wb.save("Live_Crypto_Data.xlsx")
                print("Data updated in Excel.")
    else:
        print("No data to process.")

if __name__ == "__main__":
    main()
